import pytest
import requests
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime
import os
import csv
import pandas as pd

# Configuration
BASE_URL = "http://localhost:8001"
EXCEL_FILE = "api_test_results.xlsx"
CSV_FILE = "api_test_results.csv"

def create_test_report():
    wb = Workbook()
    ws = wb.active
    ws.title = "API Test Results"
    
    # Set up headers
    headers = ["Test Case", "Endpoint", "Method", "Expected Status", "Actual Status", 
              "Expected Response", "Actual Response", "Status", "Timestamp"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
    
    return wb, ws

def create_csv_report():
    # Create CSV file with headers
    with open(CSV_FILE, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            "Test ID",
            "Test Case",
            "Endpoint",
            "Method",
            "Field",
            "Expected Value",
            "Actual Value",
            "Match Status",
            "Timestamp"
        ])

def add_to_csv(test_id, test_case, endpoint, method, response_data):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Prepare rows for CSV
    rows = []
    
    # Add status code comparison
    rows.append([
        test_id,
        test_case,
        endpoint,
        method,
        "Status Code",
        response_data["expected_status"],
        response_data["actual_status"],
        "PASS" if response_data["expected_status"] == response_data["actual_status"] else "FAIL",
        timestamp
    ])
    
    # Add response structure comparison
    if isinstance(response_data["expected_response"], dict):
        for key in response_data["expected_response"].keys():
            expected_val = response_data["expected_response"].get(key)
            actual_val = response_data["actual_response"].get(key) if isinstance(response_data["actual_response"], dict) else None
            rows.append([
                test_id,
                test_case,
                endpoint,
                method,
                f"Response.{key}",
                str(expected_val),
                str(actual_val),
                "PASS" if expected_val == actual_val else "FAIL",
                timestamp
            ])
    
    # Write to CSV
    with open(CSV_FILE, 'a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(rows)

def add_test_result(ws, row_num, test_case, endpoint, method, expected_status, 
                    actual_status, expected_response, actual_response):
    # Add to Excel
    ws.cell(row=row_num, column=1, value=test_case)
    ws.cell(row=row_num, column=2, value=endpoint)
    ws.cell(row=row_num, column=3, value=method)
    ws.cell(row=row_num, column=4, value=expected_status)
    ws.cell(row=row_num, column=5, value=actual_status)
    ws.cell(row=row_num, column=6, value=str(expected_response))
    ws.cell(row=row_num, column=7, value=str(actual_response))
    
    status = "PASS" if expected_status == actual_status else "FAIL"
    status_cell = ws.cell(row=row_num, column=8, value=status)
    
    if status == "PASS":
        status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    else:
        status_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    
    ws.cell(row=row_num, column=9, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    
    # Add to CSV
    response_data = {
        "expected_status": expected_status,
        "actual_status": actual_status,
        "expected_response": expected_response,
        "actual_response": actual_response
    }
    add_to_csv(row_num-1, test_case, endpoint, method, response_data)

def test_list_resumes_endpoint(ws, row_num):
    endpoint = "/list-resumes"
    test_case = "List Resumes"
    method = "GET"
    expected_status = 200
    expected_response = {"resumes": []}  # At minimum, should return an empty list
    
    try:
        response = requests.get(f"{BASE_URL}{endpoint}")
        actual_status = response.status_code
        actual_response = response.json()
        
        # Verify response structure
        assert "resumes" in actual_response
        assert isinstance(actual_response["resumes"], list)
        
    except Exception as e:
        actual_status = 500
        actual_response = str(e)
    
    add_test_result(ws, row_num, test_case, endpoint, method,
                    expected_status, actual_status, expected_response, actual_response)
    
    return row_num + 1

def test_rank_resumes_endpoint(ws, row_num):
    endpoint = "/rank-resumes"
    test_case = "Rank Resumes"
    method = "POST"
    expected_status = 200
    
    test_cases = [
        {
            "name": "Software Engineer Position",
            "data": {
                "job_title": "Senior Software Engineer",
                "job_description": "Looking for a Python developer with AWS experience",
                "required_skills": "Python, AWS, Docker"
            },
            "expected_response": {
                "results": []  # Should contain ranked resumes
            }
        },
        {
            "name": "Data Scientist Position",
            "data": {
                "job_title": "Data Scientist",
                "job_description": "Need an experienced data scientist with ML expertise",
                "required_skills": "Python, R, Machine Learning"
            },
            "expected_response": {
                "results": []  # Should contain ranked resumes
            }
        },
        {
            "name": "Empty Skills Test",
            "data": {
                "job_title": "Test Position",
                "job_description": "Test description",
                "required_skills": ""
            },
            "expected_response": {
                "results": []  # Should contain ranked resumes
            }
        }
    ]
    
    for test in test_cases:
        try:
            response = requests.post(
                f"{BASE_URL}{endpoint}",
                data=test["data"]
            )
            actual_status = response.status_code
            actual_response = response.json()
            
            # Verify response structure
            if actual_status == 200:
                assert "results" in actual_response
                assert isinstance(actual_response["results"], list)
                
        except Exception as e:
            actual_status = 500
            actual_response = str(e)
        
        add_test_result(ws, row_num, f"{test_case} - {test['name']}", 
                       endpoint, method, expected_status, actual_status,
                       test["expected_response"], actual_response)
        row_num += 1
    
    return row_num

def create_summary_report():
    # Read the CSV file
    df = pd.read_csv(CSV_FILE)
    
    # Create summary statistics
    summary = {
        "Total Tests": len(df["Test ID"].unique()),
        "Passed Tests": len(df[df["Match Status"] == "PASS"]["Test ID"].unique()),
        "Failed Tests": len(df[df["Match Status"] == "FAIL"]["Test ID"].unique()),
        "Total Assertions": len(df),
        "Passed Assertions": len(df[df["Match Status"] == "PASS"]),
        "Failed Assertions": len(df[df["Match Status"] == "FAIL"])
    }
    
    # Calculate pass rate
    summary["Test Pass Rate"] = f"{(summary['Passed Tests'] / summary['Total Tests'] * 100):.2f}%"
    summary["Assertion Pass Rate"] = f"{(summary['Passed Assertions'] / summary['Total Assertions'] * 100):.2f}%"
    
    # Create summary CSV
    with open("test_summary.csv", 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["Metric", "Value"])
        for key, value in summary.items():
            writer.writerow([key, value])
    
    return summary

def run_all_tests():
    print("Starting API tests...")
    
    # Create report files
    wb, ws = create_test_report()
    create_csv_report()
    row_num = 2  # Start after headers
    
    # Run tests
    row_num = test_list_resumes_endpoint(ws, row_num)
    row_num = test_rank_resumes_endpoint(ws, row_num)
    
    # Auto-adjust Excel column widths
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Save Excel results
    wb.save(EXCEL_FILE)
    
    # Create summary report
    summary = create_summary_report()
    
    # Print results
    print("\nTest Summary:")
    for metric, value in summary.items():
        print(f"{metric}: {value}")
    
    print(f"\nDetailed results have been saved to:")
    print(f"- Excel: {EXCEL_FILE}")
    print(f"- CSV: {CSV_FILE}")
    print(f"- Summary: test_summary.csv")

if __name__ == "__main__":
    run_all_tests() 